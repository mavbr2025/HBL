from pathlib import Path
from copy import copy

from openpyxl import Workbook, load_workbook

from decimal import Decimal

from mtm_hbl.excel.hbl_writer import ExcelHblWriter, ExcelWriteError

from tests.conftest import valid_data


def test_master_template_is_not_modified(tmp_path, app_config):
    template = tmp_path / "template.xlsx"
    output = tmp_path / "output.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["B4"] = "MASTER"
    worksheet["B63"] = "=SUM(J54:L54)"
    workbook.save(template)

    ExcelHblWriter(app_config).write(template, output, valid_data())

    master = load_workbook(template)
    generated = load_workbook(output)
    assert master.active["B4"].value == "MASTER"
    assert generated.active["B4"].value == "SHIPPER SA"
    assert generated.active["B5"].value == "ADDRESS"


def test_formula_cells_are_not_overwritten(tmp_path, app_config):
    template = tmp_path / "template.xlsx"
    output = tmp_path / "output.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["B4"] = "=TODAY()"
    workbook.save(template)

    try:
        ExcelHblWriter(app_config).write(template, output, valid_data())
    except ExcelWriteError as exc:
        assert "Refusing to overwrite formula cell B4" in str(exc)
    else:
        raise AssertionError("Expected ExcelWriteError")


def test_freight_rate_writes_to_charge_row_and_preserves_formula(tmp_path, app_config):
    template = tmp_path / "template.xlsx"
    output = tmp_path / "output.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["L55"] = "=D55*C72"
    workbook.save(template)
    data = valid_data()
    data.charges.charge_description = "Ocean Basic Freight"
    data.charges.unit_rate = "3000.00"
    data.charges.unit = "Per Container"
    data.charges.currency = "USD"

    ExcelHblWriter(app_config).write(template, output, data)

    generated = load_workbook(output, data_only=False)
    assert generated.active["B55"].value == "Ocean Basic Freight"
    assert generated.active["D55"].value == "3000.00"
    assert generated.active["F55"].value == "Per Container"
    assert generated.active["H55"].value == "USD"
    assert generated.active["L55"].value == "=D55*C72"


def test_container_weights_are_numeric_for_excel_totals(tmp_path, app_config):
    template = tmp_path / "template.xlsx"
    output = tmp_path / "output.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    workbook.save(template)

    ExcelHblWriter(app_config).write(template, output, valid_data())

    generated = load_workbook(output, data_only=False)
    assert generated.active["J37"].value == Decimal("1000.00")
    assert generated.active["L37"].value == Decimal("12.50")
    assert generated.active["J62"].value == "DRAFT"


def test_container_marks_align_with_weight_rows(tmp_path, app_config):
    template = tmp_path / "template.xlsx"
    output = tmp_path / "output.xlsx"
    workbook = Workbook()
    workbook.save(template)
    data = valid_data()
    data.containers[0].container_no = "CONT1"
    data.containers[0].seal_no = "SEAL1"
    for index in range(2, 5):
        container = data.containers[0].model_copy(deep=True)
        container.container_no = f"CONT{index}"
        container.seal_no = f"SEAL{index}"
        container.gross_weight = str(index * 1000)
        data.containers.append(container)

    ExcelHblWriter(app_config).write(template, output, data)

    generated = load_workbook(output, data_only=False)
    sheet = generated.active
    assert sheet["C37"].value == "CONT1"
    assert sheet["C39"].value == "CONT2"
    assert sheet["C41"].value == "CONT3"
    assert sheet["C43"].value == "CONT4"
    assert sheet["J37"].value == Decimal("1000.00")
    assert sheet["J39"].value == Decimal("2000")
    assert sheet["J41"].value == Decimal("3000")
    assert sheet["J43"].value == Decimal("4000")


def test_cargo_description_keeps_document_body_font_size(tmp_path, app_config):
    template = tmp_path / "template.xlsx"
    output = tmp_path / "output.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    font = copy(worksheet["D32"].font)
    font.sz = 6
    worksheet["D32"].font = font
    workbook.save(template)
    data = valid_data()
    data.cargo.description_raw = "\n".join(
        [
            "SHIPPER'S LOAD,COUNT & SEAL",
            "4 containers said to contain 21 SKIDS",
            "CONTAINER: MRKU0931970",
        ]
    )

    ExcelHblWriter(app_config).write(template, output, data)

    generated = load_workbook(output, data_only=False)
    target = generated.active["D32"]
    assert target.value == data.cargo.description_raw
    assert target.font.sz == 9
    assert target.alignment.wrap_text is True
    assert target.alignment.shrink_to_fit is not True


def test_delivery_apply_to_renders_consignee_block(tmp_path, app_config):
    template = tmp_path / "template.xlsx"
    output = tmp_path / "output.xlsx"
    workbook = Workbook()
    workbook.save(template)
    data = valid_data()
    data.parties.consignee.raw_text = "CONSIGNEE SA\nCONSIGNEE ADDRESS"
    data.parties.delivery_apply_to.raw_text = "MTM DELIVERY BLOCK"

    ExcelHblWriter(app_config).write(template, output, data)

    generated = load_workbook(output, data_only=False)
    assert generated.active["G24"].value == "CONSIGNEE SA"
    assert generated.active["G25"].value == "CONSIGNEE ADDRESS"


def test_entity_header_is_written_below_logo(tmp_path, app_config):
    template = tmp_path / "template.xlsx"
    output = tmp_path / "output.xlsx"
    workbook = Workbook()
    workbook.save(template)

    ExcelHblWriter(app_config).write(template, output, valid_data())

    generated = load_workbook(output, data_only=False)
    sheet = generated.active
    assert sheet["H17"].value == "MTM LOGIX GUATEMALA SOCIEDAD ANONIMA"
    assert sheet["H18"].value == "NIT: 109582985"
    assert sheet["H19"].value == "7A. AV. 131-78 ZONA 4"
    assert sheet["H20"].value == "EDIFICIO SEPTIMO, NIVEL 3, OFICINA 306"


def test_port_of_discharge_is_left_aligned(tmp_path, app_config):
    template = tmp_path / "template.xlsx"
    output = tmp_path / "output.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    alignment = copy(worksheet["A28"].alignment)
    alignment.horizontal = "right"
    worksheet["A28"].alignment = alignment
    workbook.save(template)

    ExcelHblWriter(app_config).write(template, output, valid_data())

    generated = load_workbook(output, data_only=False)
    assert generated.active["A28"].alignment.horizontal == "left"
