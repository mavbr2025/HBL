from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def inspect_template(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=False)
    sheets = []
    for sheet in workbook.worksheets:
        sheets.append(
            {
                "title": sheet.title,
                "sheet_state": sheet.sheet_state,
                "merged_ranges": [str(item) for item in sheet.merged_cells.ranges],
                "print_area": sheet.print_area,
                "protected": bool(sheet.protection.sheet),
                "formula_cells": [
                    cell.coordinate
                    for row in sheet.iter_rows()
                    for cell in row
                    if isinstance(cell.value, str) and cell.value.startswith("=")
                ],
            }
        )
    return {"sheets": sheets}
