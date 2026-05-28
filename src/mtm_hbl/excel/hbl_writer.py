import shutil
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter
from copy import copy
from pydantic import BaseModel

from mtm_hbl.config import AppConfig
from mtm_hbl.models.canonical import CanonicalHblData
from mtm_hbl.utils.values import get_path


class ExcelWriteError(RuntimeError):
    pass


class ExcelHblWriter:
    def __init__(self, app_config: AppConfig) -> None:
        self.app_config = app_config
        self.mapping = app_config.excel_cell_mapping

    def write(self, template_path: Path, output_path: Path, data: CanonicalHblData) -> Path:
        if not template_path.exists():
            raise FileNotFoundError(f"Template file not found: {template_path}")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(template_path, output_path)

        workbook = load_workbook(output_path)
        sheet = workbook.active
        merged_top_lefts = {
            str(range_.start_cell.coordinate): str(range_)
            for range_ in sheet.merged_cells.ranges
        }

        for key, spec in self.mapping.get("cells", {}).items():
            if data.charges.line_items:
                if key == "freight_and_charges":
                    self._write_charge_line_items(sheet, data)
                    continue
                if key in {
                    "freight_rate",
                    "freight_unit",
                    "freight_currency",
                    "prepaid_charges",
                    "collect_charges",
                }:
                    continue
            cell = spec.get("cell")
            if not cell or not spec.get("writable"):
                continue
            if spec.get("formula_expected"):
                continue
            if key in {"container_marks_start", "gross_weight_start", "measurement_start"}:
                self._write_container_values(sheet, key, cell, data)
                continue
            if "value" in spec:
                value = spec["value"]
            else:
                value = self._format_value(key, get_path(data, spec.get("source_path", "")))
            if value is None:
                continue
            if key == "cargo_description":
                self._write_cargo_description_block(sheet, cell, str(value))
                continue
            if key in {"shipper", "consignee", "notify_party", "delivery_apply_to"}:
                self._write_multiline_block(sheet, cell, str(value))
                continue
            target = sheet[cell]
            if isinstance(target.value, str) and target.value.startswith("="):
                raise ExcelWriteError(f"Refusing to overwrite formula cell {cell} for mapping {key}.")
            if cell not in merged_top_lefts:
                for merged_range in sheet.merged_cells.ranges:
                    if cell in merged_range and cell != merged_range.start_cell.coordinate:
                        raise ExcelWriteError(
                            f"Cell {cell} is inside merged range {merged_range}; configure top-left cell."
                        )
            target.value = value
            if isinstance(value, str) and "\n" in value:
                alignment = copy(target.alignment)
                alignment.wrap_text = True
                target.alignment = alignment
            if key in {"hbl_number", "mbl_number"}:
                alignment = copy(target.alignment)
                alignment.shrink_to_fit = True
                target.alignment = alignment
            if key in {
                "place_of_receipt",
                "port_of_loading",
                "port_of_discharge",
                "place_of_delivery",
                "vessel_voyage",
            }:
                self._style_routing_value_cell(target, key, value)

        self._write_entity_header_below_logo(sheet, data)
        self._fit_to_single_page(sheet)
        if sheet["J52"].data_type == "f":
            sheet["J52"].number_format = '#,##0.000\\ "KGS"'
        workbook.save(output_path)
        return output_path

    def _write_multiline_block(self, sheet, start_cell: str, value: str) -> None:
        start_row, start_col = coordinate_to_tuple(start_cell)
        lines = [line.strip() for line in value.splitlines() if line.strip()]
        if not lines:
            sheet[start_cell].value = ""
            return
        for index, line in enumerate(lines):
            row = start_row + index
            coordinate = f"{get_column_letter(start_col)}{row}"
            target = sheet[coordinate]
            if isinstance(target, MergedCell) or (index > 0 and target.value not in (None, "")):
                previous = sheet[f"{get_column_letter(start_col)}{row - 1}"]
                remaining = " ".join(lines[index:])
                previous.value = f"{previous.value} {remaining}".strip()
                break
            if isinstance(target.value, str) and target.value.startswith("="):
                raise ExcelWriteError(f"Refusing to overwrite formula cell {coordinate}.")
            if index > 0:
                source = sheet[start_cell]
                if target.value in (None, ""):
                    target._style = copy(source._style)
                    target.font = copy(source.font)
                    target.alignment = copy(source.alignment)
                    target.border = copy(source.border)
                    target.fill = copy(source.fill)
            target.value = line
            if len(lines) > 4 or len(line) > 45:
                font = copy(target.font)
                font.sz = 9
                target.font = font

    @staticmethod
    def _style_routing_value_cell(target, key: str, value: object) -> None:
        text = str(value or "")

        alignment = copy(target.alignment)
        alignment.wrap_text = False
        alignment.shrink_to_fit = len(text) > 36
        if key == "port_of_discharge":
            alignment.horizontal = "left"
        target.alignment = alignment

        font = copy(target.font)
        if key == "vessel_voyage":
            font.sz = 9
        elif len(text) > 24:
            font.sz = 7
        else:
            font.sz = 8
        target.font = font

    @staticmethod
    def _write_cargo_description_block(sheet, cell: str, value: str) -> None:
        target = sheet[cell]
        if isinstance(target.value, str) and target.value.startswith("="):
            raise ExcelWriteError(f"Refusing to overwrite formula cell {cell}.")
        target.value = value

        font = copy(target.font)
        font.sz = max(float(font.sz or 0), 9)
        target.font = font

        alignment = copy(target.alignment)
        alignment.wrap_text = True
        alignment.shrink_to_fit = False
        target.alignment = alignment

    def _write_entity_header_below_logo(self, sheet, data: CanonicalHblData) -> None:
        country = data.scope.owner_country or data.scope.country
        entity = self.app_config.entity_rules.get("entity_by_owner_country", {}).get(country, {})
        header = entity.get("header_below_logo", {})
        start_cell = header.get("start_cell")
        lines = [str(line) for line in header.get("lines", []) if str(line).strip()]
        if not start_cell or not lines:
            return

        start_row, start_col = coordinate_to_tuple(start_cell)
        for offset, line in enumerate(lines):
            target = sheet[f"{get_column_letter(start_col)}{start_row + offset}"]
            target.value = line
            font = copy(target.font)
            font.sz = 9
            target.font = font
            alignment = copy(target.alignment)
            alignment.wrap_text = False
            alignment.shrink_to_fit = True
            target.alignment = alignment

    def _write_container_values(self, sheet, key: str, start_cell: str, data: CanonicalHblData) -> None:
        if key == "container_marks_start":
            self._write_container_marks_block(sheet, data)
            return
        start_row, start_col = coordinate_to_tuple(start_cell)
        for index, container in enumerate(data.containers[:5]):
            row = start_row + (index * 2)
            coordinate = f"{get_column_letter(start_col)}{row}"
            target = sheet[coordinate]
            if isinstance(target.value, str) and target.value.startswith("="):
                raise ExcelWriteError(f"Refusing to overwrite formula cell {coordinate} for {key}.")
            if key == "gross_weight_start":
                self._copy_container_measure_style(sheet, start_row, row, start_col)
                value = self._number_for_excel(container.gross_weight)
                target.value = value
                target.number_format = '#,##0.000'
                if value != "":
                    self._write_unit_label(sheet, row, start_col + 1, container.gross_weight_unit, start_row)
            elif key == "measurement_start":
                self._copy_container_measure_style(sheet, start_row, row, start_col)
                value = self._number_for_excel(container.measurement)
                target.value = value
                target.number_format = '#,##0.000'
                if value != "":
                    self._write_unit_label(sheet, row, start_col + 1, container.measurement_unit, start_row)
                else:
                    sheet[f"{get_column_letter(start_col + 1)}{row}"].value = None

    def _write_container_marks_block(self, sheet, data: CanonicalHblData) -> None:
        for row in range(33, 51):
            for column in ["B", "C"]:
                sheet[f"{column}{row}"].value = None
        sheet["B32"].value = "MARKS & NOS.:"

        marks = [
            line.strip()
            for line in (data.containers[0].marks_and_numbers if data.containers else "").splitlines()
            if line.strip()
        ]
        for offset, line in enumerate(marks[:3]):
            self._write_small_mark_cell(sheet[f"B{33 + offset}"], line)

        for offset, container in enumerate(data.containers[:5]):
            row = 37 + (offset * 2)
            self._write_small_mark_cell(sheet[f"B{row}"], "CONTAINER:")
            self._write_small_mark_cell(sheet[f"C{row}"], container.container_no)
            self._write_small_mark_cell(sheet[f"B{row + 1}"], "SEAL:")
            self._write_small_mark_cell(sheet[f"C{row + 1}"], container.seal_no)

    @staticmethod
    def _write_small_mark_cell(target, value: str) -> None:
        target.value = value
        font = copy(target.font)
        font.sz = 7
        target.font = font
        alignment = copy(target.alignment)
        alignment.wrap_text = False
        alignment.shrink_to_fit = True
        target.alignment = alignment

    def _write_charge_line_items(self, sheet, data: CanonicalHblData) -> None:
        start_row = 55
        max_rows = 7
        for row in range(start_row, start_row + max_rows):
            for column in ["B", "D", "F", "H", "J", "L"]:
                if row > start_row:
                    self._copy_cell_style(sheet[f"{column}{row - 1}"], sheet[f"{column}{row}"])
                sheet[f"{column}{row}"].value = None
        for offset, item in enumerate(data.charges.line_items[:max_rows]):
            row = start_row + offset
            sheet[f"B{row}"].value = item.description
            sheet[f"D{row}"].value = self._number_for_excel(item.rate)
            sheet[f"D{row}"].number_format = '#,##0.00'
            rate_alignment = copy(sheet[f"D{row}"].alignment)
            rate_alignment.shrink_to_fit = True
            sheet[f"D{row}"].alignment = rate_alignment
            sheet[f"F{row}"].value = item.unit
            sheet[f"H{row}"].value = item.currency
            if item.prepaid_amount:
                sheet[f"J{row}"].value = self._number_for_excel(item.prepaid_amount)
                sheet[f"J{row}"].number_format = '#,##0.00'
            if item.collect_amount:
                sheet[f"L{row}"].value = self._number_for_excel(item.collect_amount)
                sheet[f"L{row}"].number_format = '#,##0.00'
        if len(data.charges.line_items) > 6 and sheet["B65"].data_type == "f":
            sheet["B65"].value = "=SUM(J55:J61)+SUM(L55:L61)"

    def _copy_container_measure_style(self, sheet, start_row: int, row: int, start_col: int) -> None:
        if row == start_row:
            return
        source = sheet[f"{get_column_letter(start_col)}{start_row}"]
        target = sheet[f"{get_column_letter(start_col)}{row}"]
        self._copy_cell_style(source, target)

    def _write_unit_label(self, sheet, row: int, col: int, value: str, start_row: int) -> None:
        if not value:
            return
        target = sheet[f"{get_column_letter(col)}{row}"]
        if row != start_row:
            self._copy_cell_style(sheet[f"{get_column_letter(col)}{start_row}"], target)
        target.value = value

    @staticmethod
    def _copy_cell_style(source, target) -> None:
        target._style = copy(source._style)
        target.font = copy(source.font)
        target.alignment = copy(source.alignment)
        target.border = copy(source.border)
        target.fill = copy(source.fill)

    @staticmethod
    def _fit_to_single_page(sheet) -> None:
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 1
        sheet.page_setup.scale = None

    @staticmethod
    def _format_value(key: str, value: object) -> object:
        if key == "container_marks_start" and isinstance(value, list):
            return "\n".join(
                "\n".join(
                    part
                    for part in [
                        getattr(container, "container_no", ""),
                        f"SEAL: {getattr(container, 'seal_no', '')}"
                        if getattr(container, "seal_no", "")
                        else "",
                        getattr(container, "marks_and_numbers", ""),
                    ]
                    if part
                )
                for container in value
            )
        if key == "gross_weight_start" and isinstance(value, list):
            return "\n".join(
                " ".join(
                    part
                    for part in [
                        getattr(container, "gross_weight", ""),
                        getattr(container, "gross_weight_unit", ""),
                    ]
                    if part
                )
                for container in value
            )
        if key == "measurement_start" and isinstance(value, list):
            return "\n".join(
                " ".join(
                    part
                    for part in [
                        getattr(container, "measurement", ""),
                        getattr(container, "measurement_unit", ""),
                    ]
                    if part
                )
                for container in value
            )
        if isinstance(value, BaseModel):
            return None
        return value

    @staticmethod
    def _number_for_excel(value: str) -> object:
        if not value:
            return ""
        cleaned = str(value).replace(",", "").strip()
        try:
            return Decimal(cleaned)
        except InvalidOperation:
            return value
